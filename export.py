from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

def create_forecast_spreadsheet(products, operating_expenses, cogs_percentage, loan_details, seasonality_factors, company_name, depreciation, interest_expense):
    """
    Creates an Excel spreadsheet with financial forecast and loan amortization data.
    """
    if seasonality_factors is None:
        seasonality_factors = [1.0] * 12

    wb = Workbook()

    # --- Sheet 1: Financial Forecast ---
    ws_revenue = wb.active
    ws_revenue.title = "Quarterly Revenue"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_font = Font(size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid") # Dark Blue
    currency_format = '$#,##0.00'

    # Title
    display_company_name = company_name if company_name else 'My Awesome Startup'
    ws_revenue['A1'] = f'Financial Forecast for {display_company_name}'
    ws_revenue['A1'].font = title_font
    ws_revenue['A1'].fill = title_fill
    
    # Create headers: Quarter | Product 1 | Product 2 | ... | Total Revenue
    product_names = [p.get('description', 'N/A') for p in products]
    headers = ['Quarter'] + product_names + ['Total Revenue']
    ws_revenue.append(headers)
    for cell in ws_revenue[2]: # Style header row
        cell.font = header_font
        cell.fill = header_fill

    # Normalize seasonality factors so their sum is 12 (average is 1)
    total_factor = sum(seasonality_factors)
    normalized_factors = [(f / total_factor) * 12 for f in seasonality_factors] if total_factor > 0 else [1.0] * 12

    # Calculate base monthly revenue for each product
    product_monthly_revenues = []
    for product in products:
        price = float(product.get('price', 0) or 0)
        sales_volume = int(product.get('sales_volume', 0) or 0)
        unit = product.get('sales_volume_unit', 'monthly')
        annual_volume = sales_volume * 12 if unit == 'monthly' else sales_volume * 4
        base_monthly_revenue = (price * annual_volume) / 12
        product_monthly_revenues.append(base_monthly_revenue)

    # Populate quarterly data
    for q in range(4): # For Q1, Q2, Q3, Q4
        quarter_data = [f'Q{q+1}']
        quarterly_total_revenue = 0
        
        for prod_idx, base_monthly_rev in enumerate(product_monthly_revenues):
            product_quarterly_revenue = 0
            for month_in_quarter in range(3):
                month_idx = q * 3 + month_in_quarter
                # Revenue for this product for this month
                product_quarterly_revenue += base_monthly_rev * normalized_factors[month_idx]
            
            quarter_data.append(product_quarterly_revenue)
            quarterly_total_revenue += product_quarterly_revenue

        quarter_data.append(quarterly_total_revenue)
        ws_revenue.append(quarter_data)

    # Apply currency formatting to all revenue cells
    for row in ws_revenue.iter_rows(min_row=3, max_row=ws_revenue.max_row, min_col=2, max_col=ws_revenue.max_column):
        for cell in row:
            cell.number_format = currency_format

    # --- Sheet 2: Annual P&L Summary ---
    ws_pnl = wb.create_sheet(title="Annual P&L Summary")
    ws_pnl['A1'] = 'Profit & Loss Summary (USD)'
    ws_pnl['A1'].font = title_font
    ws_pnl['A1'].fill = title_fill
    ws_pnl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    pnl_headers = [
        'Year', 'Total Revenue', 'COGS', 'Gross Profit', 'Operating Expenses',
        'Net Operating Income', 'Depreciation', 'Earnings Before Tax', 'Taxes',
        'Net Income', 'Debt Service Coverage Ratio'
    ]
    ws_pnl.append(pnl_headers)
    for cell in ws_pnl[2]:
        cell.font = header_font
        cell.fill = header_fill

    # P&L Calculation
    # Year 1 figures
    y1_revenue = sum(product_monthly_revenues) * 12
    y1_opex = 0
    for expense in operating_expenses:
        amount = float(expense.get('amount', 0))
        if expense.get('frequency') == 'monthly':
            y1_opex += amount * 12
        elif expense.get('frequency') == 'quarterly':
            y1_opex += amount * 4

    # Growth assumptions
    revenue_growth_rate = 0.10  # 10%
    opex_growth_rate = 0.05     # 5%

    current_revenue = y1_revenue
    current_opex = y1_opex

    for year in range(1, 6):
        if year > 1:
            current_revenue *= (1 + revenue_growth_rate)
            current_opex *= (1 + opex_growth_rate)

        cogs = current_revenue * (cogs_percentage / 100)
        gross_profit = current_revenue - cogs
        net_operating_income = gross_profit - current_opex
        # EBT calculation in this app seems to be NOI - Depreciation - Interest
        ebt = net_operating_income - depreciation - interest_expense
        
        # Assuming a simple tax rate calculation for projection
        tax_rate_from_loan_calc = 0.25 # A standard assumption for projections
        taxes = max(0, ebt * tax_rate_from_loan_calc)
        net_income = ebt - taxes

        # DSCR Calculation
        total_debt_service = (loan_details.get('monthly_payment', 0) or 0) * 12
        dscr = (net_operating_income / total_debt_service) if total_debt_service > 0 else 0

        row_data = [
            year, current_revenue, cogs, gross_profit, current_opex,
            net_operating_income, depreciation, ebt, taxes, net_income,
            dscr if dscr > 0 else 'N/A'
        ]
        ws_pnl.append(row_data)

    # Formatting for P&L sheet
    for row in ws_pnl.iter_rows(min_row=3, max_row=ws_pnl.max_row, min_col=2, max_col=10):
        for cell in row:
            cell.number_format = currency_format
    
    # Format DSCR column
    for cell in ws_pnl['K']:
        if cell.row > 2:
            cell.number_format = '0.00'


    # --- Sheet 2: Loan Amortization ---
    if loan_details and loan_details.get('schedule'):
        ws2 = wb.create_sheet(title="Loan Amortization") # This will now be sheet 3

        # Loan Summary
        ws2['A1'] = 'Loan Summary'
        ws2['A1'].font = title_font
        ws2.append(['Loan Amount', loan_details.get('loan_amount')])
        ws2.append(['Annual Interest Rate (%)', loan_details.get('interest_rate')])
        ws2.append(['Loan Term (Years)', loan_details.get('loan_term')])
        ws2.append(['Monthly Payment', loan_details.get('monthly_payment')])
        
        # Apply formatting to summary
        for row_num in range(2, 6):
            ws2[f'B{row_num}'].number_format = currency_format if row_num in [2, 5] else '0.00'

        # Amortization Schedule Headers
        schedule_headers = ['Month', 'Principal Payment', 'Interest Payment', 'Remaining Balance']
        ws2.append([]) # Spacer row
        ws2.append(schedule_headers)
        for cell in ws2[7]:
            cell.font = header_font
            cell.fill = header_fill

        # Schedule Data
        for item in loan_details['schedule']:
            ws2.append([
                item['month'],
                item['principal_payment'],
                item['interest_payment'],
                item['remaining_balance']
            ])
        
        # Apply currency formatting to schedule
        for row in ws2.iter_rows(min_row=8, max_row=ws2.max_row, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = currency_format

    # Adjust column widths for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
    
    # Merge the title cell after calculating column widths
    ws_revenue.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws_revenue.max_column)

    # Save to an in-memory file
    in_memory_file = BytesIO()
    wb.save(in_memory_file)
    in_memory_file.seek(0)
    return in_memory_file